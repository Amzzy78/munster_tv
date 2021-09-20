import openpyxl
import gspread
from google.oauth2.service_account import Credentials



SCOPE = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.file",
    "https://www.googleapis.com/auth/drive"
    ] 

CREDS = Credentials.from_service_account_file('creds.json')
SCOPED_CREDS = CREDS.with_scopes(SCOPE)
GSPREAD_CLIENT = gspread.authorize(SCOPED_CREDS)

SHEET = GSPREAD_CLIENT.open('munster-tv')
# Balance sheet
balance = SHEET.worksheet('balance')

data = balance.get_all_values()

#print(data)

# Update Inventory sheet
inv_file = openpyxl.load_workbook('munster-tv')
# Variable to work on inventory sheet
product_list = inv_file['inventory'] 

# Calculate how many products per supplier
product_per_supplier = {} # Dict
# Loop through the rows
for product_row in  int(product_list.max_row)

print(product_list.max_row)
